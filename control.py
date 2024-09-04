import random
import time
from tkinter import filedialog
import string
from openpyxl import load_workbook
import xlrd
import re
from decimal import Decimal, getcontext
from tkinter import messagebox
import json
import math

class Controller:
    ui: object
    def __init__(self):
        self.load_if = None
        pass
    def init(self, ui):
        self.ui = ui

    def save_txt(self):
    # 弹出文件保存对话框
        file_path = filedialog.asksaveasfilename(defaultextension='.txt',
                                                filetypes=[('Text Files', '*.txt')],
                                                title='保存为TXT文件')
        # 如果用户取消了保存，则直接返回
        if not file_path:
            return
        try:
            # 打开选定的文件路径，以写入模式保存数据
            with open(file_path, 'w', encoding='utf-8') as f:
                # 获取表格数据，这里假设表格数据是字符串格式
                for row in self.ui.tk_table_num_collect.get_children():
                    row_data = self.ui.tk_table_num_collect.item(row)['values']
                # 将每行数据转换为字符串并写入文件
                    row_str = '\t'.join(str(item) for item in row_data) + '\n'
                    f.write(row_str)
        except Exception as e:
            return

    def limit_number(self,value, max_value):
        return min(value, max_value)

    def generate_numbers_from_range(self, start_num, end_num, generation_num, repeat_if,float_precision,load_num=None, ):
        if float_precision != 0:
            return self._generate_floats(start_num, end_num, generation_num, repeat_if, float_precision)
        else:
            return self._generate_integers(start_num, end_num, generation_num, repeat_if, load_num)

    def _generate_floats(self, start_num, end_num, generation_num, repeat_if, float_precision):
        # 设置 Decimal 的精度
        getcontext().prec = float_precision + 5  # 增加额外的精度以防计算误差

        def generate_float():
            # 使用 Decimal 生成精确的浮点数
            return round(Decimal(random.uniform(start_num, end_num)), float_precision)

        if repeat_if == "重复":
            return [generate_float() for _ in range(generation_num)]

        # 对于不重复的浮点数生成
        # 计算可能的浮点数范围
        step = Decimal(1) / (10 ** float_precision)
        num_steps = int((Decimal(end_num) - Decimal(start_num)) / step) + 1

        if generation_num > num_steps:
            generation_num = num_steps

        # 生成不重复的浮点数
        float_values = set()
        while len(float_values) < generation_num:
            value = generate_float()
            if Decimal(start_num) <= value <= Decimal(end_num):
                float_values.add(value)

        return list(float_values)

    def _generate_integers(self, start_num, end_num, generation_num, repeat_if, load_num):
        def handle_repeating_logic(start, end, num, repeat_mode, max_num=None):
            if max_num is not None:
                if max_num < start:
                    return []
                end = min(end, max_num)
            if repeat_mode == "重复":
                return [random.randint(start, end) for _ in range(num)]
            elif repeat_mode == "不重复":
                available_range_size = end - start + 1
                if num > available_range_size:
                    num = available_range_size
                return random.sample(range(start, end + 1), num)

        if load_num is not None:
            return handle_repeating_logic(start_num, end_num, generation_num, repeat_if, load_num)

        return handle_repeating_logic(start_num, end_num, generation_num, repeat_if)

    def validate_password_strength(self, password):
        # 使用正则表达式检查密码是否包含大写字母、小写字母、数字和特殊字符
        has_upper = bool(re.search(r'[A-Z]', password))
        has_lower = bool(re.search(r'[a-z]', password))
        has_digit = bool(re.search(r'\d', password))
        has_special = bool(re.search(r'[!@#%^&*(),.?":{}|<>]', password))

        # 根据符合的条件数量，判断密码强度
        strength = sum([has_upper, has_lower, has_digit, has_special])

        if strength == 4:
            return "强密码"
        elif strength == 3:
            return "中密码"
        else:
            return "低密码"

    def create_number(self, start_num, end_num, generation_num, repeat_if, method, clear_if, special_setting):
        if method == "密码":
            remove_least = special_setting.get("password", {}).get("remove_least", False)
            if remove_least:
                num_true_settings = sum([
                    special_setting.get("password", {}).get("include_numbers",True),
                    special_setting.get("password", {}).get("include_lower",True),
                    special_setting.get("password", {}).get("include_character",True),
                    special_setting.get("password", {}).get("include_special_chars",True)
                ])
                start_num = max(1, num_true_settings, start_num)
                end_num = max(1, num_true_settings, end_num, start_num)
            else:
                start_num = max(10, start_num)
                end_num = max(10, end_num, start_num)
            self.ui.tk_input_start_num.delete(0, "end")
            self.ui.tk_input_start_num.insert(0, start_num)
            self.ui.tk_input_end_num.delete(0, "end")
            self.ui.tk_input_end_num.insert(0, end_num)
            generated_numbers = self.generate_random_passwords(start_num, end_num, generation_num,special_setting.get("password", {
                "include_numbers": True,
                "include_lower": True,
                "include_character": True,
                "include_special_chars": True,
                "exclude_similar_chars": False}))
        else:
            generate_content = special_setting.get("random_settings", {}).get("generate_content", {})
            other_setting = special_setting.get("random_settings", {})
            json_content = {}
            if generate_content:
                json_content = json.loads(generate_content)

            if start_num > end_num:
                start_num, end_num = end_num, start_num
            if method == "字母":
                letter_case = special_setting.get("letter", {"letter_case": "lower"}).get("letter_case", "lower")
                characters_range = self.get_characters_range(letter_case, start_num, end_num)
                generated_numbers = self.generate_numbers_from_characters(generation_num, repeat_if, characters_range)
            elif method in json_content:
                generated_numbers = self.generate_custom_content(json_content[method], generation_num, other_setting)
            elif method not in ["数字", "字母", "密码", "自定义"] and self.load_if:
                start_num = max(1, start_num)
                end_num = max(1, end_num)
                names_from_file = self.handle_file_data(method, end_num, self.load_if)
                generated_indices = self.generate_numbers_from_range(start_num, end_num, generation_num, repeat_if, 0, load_num=len(names_from_file))
                generated_numbers = [names_from_file[i-1] for i in generated_indices]
            elif method == "自定义":
                generated_numbers = self.generate_custom_content(generate_content, generation_num,other_setting)
            else:
                generated_numbers = self.generate_numbers_from_range(start_num, end_num, generation_num, repeat_if,float_precision=special_setting.get("number", {"float_precision": 0}).get("float_precision", 0))

        if clear_if == "自动清除":
            self.ui.tk_table_num_collect.delete(*self.ui.tk_table_num_collect.get_children())

        self.ui.tk_button_create_num.config(state="disabled")

        for i, num in enumerate(generated_numbers, start=1):
            display_value = num
            if method == "密码":
                password_strength = self.validate_password_strength(num)
                self.ui.tk_label_now_num.config(text=password_strength)
            else:
                self.ui.tk_label_now_num.config(text=display_value)
            self.ui.update()
            time.sleep(0.01)
            self.ui.tk_table_num_collect.insert('', 'end', values=(i, display_value))
        self.ui.tk_button_create_num.config(state="normal")

    def generate_custom_content(self, generate_content, generation_num, special_setting):
        append_unit = special_setting.get("append_unit", "")
        append_prefix = special_setting.get("append_prefix", "")
        append_interval = special_setting.get("append_interval", "")
        enable_cross_generation = special_setting.get("enable_cross_generation", False)
        enable_random_cross = special_setting.get("enable_random_cross", False)
        distribution_type = special_setting.get("distribution_type", 'uniform')

        if isinstance(generate_content, str):
            try:
                generate_content = json.loads(generate_content)
            except json.JSONDecodeError:
                if generate_content == '':
                    messagebox.showerror("错误", "生成内容不能为空！")
                else:
                    messagebox.showerror("错误", "生成内容格式错误,无法解析为JSON")
                return []
        elif isinstance(generate_content, list):
            results = []
            for _ in range(generation_num):
                result = random.choice(generate_content)
                if append_unit:
                    result += " " + append_unit
                if append_prefix:
                    result = append_prefix + " " + result
                results.append(result)
            return results
        elif not isinstance(generate_content, dict):
            messagebox.showerror("错误", "生成内容格式错误，必须是字典、列表或可解析为字典的字符串")
            return []

        categories = list(generate_content.keys())
        results = []

        if distribution_type == 'normal':
            mean = len(categories) / 2
            std_dev = len(categories) / 6
            category_weights = [math.exp(-0.5 * ((i - mean) / std_dev) ** 2) for i in range(len(categories))]
            category_total = sum(category_weights)
            category_weights = [w / category_total for w in category_weights]
            category_probs = dict(zip(categories, category_weights))

        for _ in range(generation_num):
            if enable_cross_generation:
                selected_items = []
                for cat in categories:
                    if distribution_type == 'normal':
                        subcategory_weights = [category_probs[cat]] * len(generate_content[cat])
                        subcategory_total = sum(subcategory_weights)
                        subcategory_weights = [w / subcategory_total for w in subcategory_weights]
                        selected_item = random.choices(generate_content[cat], weights=subcategory_weights)[0]
                    else:
                        selected_item = random.choice(generate_content[cat])
                    selected_items.append(selected_item)

                if enable_random_cross:
                    random.shuffle(selected_items)

                result = append_interval.join(selected_items)
            else:
                if distribution_type == 'normal':
                    category = random.choices(categories, weights=[category_probs.get(cat, 1.0) for cat in categories])[0]
                else:
                    category = random.choice(categories)
                result = random.choice(generate_content[category])

            if append_unit:
                result += " " + append_unit

            if append_prefix:
                result = append_prefix + " " + result

            results.append(result)

        return results

    def generate_random_passwords(self, min_length, max_length, generation_num, settings):
        include_numbers = settings.get("include_numbers", True)
        include_lowers = settings.get("include_lower", True)
        include_uppers = settings.get("include_character", True)
        include_special_chars = settings.get("include_special_chars", True)
        exclude_similar_chars = settings.get("exclude_similar_chars", False)

        # 初始化字符集
        lowers = string.ascii_lowercase if include_lowers else ""
        uppers = string.ascii_uppercase if include_uppers else ""
        numbers = string.digits if include_numbers else ""
        special_chars = "!@#%^&*" if include_special_chars else ""
        characters = lowers + uppers + numbers + special_chars

        num_true_settings = sum([
            include_numbers,
            include_lowers,
            include_uppers,
            include_special_chars
        ])

        # 处理排除相似字符的设置
        if exclude_similar_chars:
            similar_chars = 'il1Lo0O'
            characters = ''.join(c for c in characters if c not in similar_chars)

        if not characters:
            messagebox.showerror("错误", "字符集为空，请检查设置")
            return []

        passwords = []
        max_attempts = 100
        for _ in range(generation_num):
            for attempt in range(max_attempts):
                password_length = random.randint(min_length, max_length)
                password = ''.join(random.choice(characters) for _ in range(password_length))

                if exclude_similar_chars:# 检查是否包含至少一个大写字母、小写字母、数字和特殊符号
                    if password_length >= num_true_settings:
                        def has_consecutive_sequence(s):
                            for i in range(len(s) - 2):
                                if ord(s[i+1]) == ord(s[i]) + 1 == ord(s[i+2]) - 1:
                                    return True
                            return False

                        categories = [
                            include_uppers and any(c in uppers for c in password),
                            include_lowers and any(c in lowers for c in password),
                            include_numbers and any(c in numbers for c in password),
                            include_special_chars and any(c in special_chars for c in password)
                        ]
                        if sum(categories) >= num_true_settings and not has_consecutive_sequence(password):
                            passwords.append(password)
                            break
                    else:
                        messagebox.showerror("错误", "增强密码的长度不得小于指定类别数！")
                        break
                else:
                    passwords.append(password)
                    break

        return passwords

    def get_characters_range(self, letter_case, start_num, end_num):
        if letter_case == "lower":
            characters = string.ascii_lowercase
        elif letter_case == "upper":
            characters = string.ascii_uppercase
        elif letter_case == "mixed":
            characters = string.ascii_letters  # 组合大小写字母
        else:
            return 'X'
        # 确保 end_num 不超过字符集长度
        if end_num > len(characters):
            end_num = len(characters)
        # 处理 start_num 可能小于 1 的情况
        if start_num < 1:
            start_num = 1
        return characters[start_num-1:end_num]

    def generate_numbers_from_characters(self, generation_num, repeat_if, characters_range):
        available_range_size = len(characters_range)
        if repeat_if == "不重复" and generation_num > available_range_size:
            generation_num = available_range_size
        if repeat_if == "重复":
            return [random.choice(characters_range) for _ in range(generation_num)]
        elif repeat_if == "不重复":
            return random.sample(characters_range, generation_num)

    def clear_table(self,evt):
        self.ui.tk_table_num_collect.delete(*self.ui.tk_table_num_collect.get_children())

    def load_execl(self, evt):
        if not self.load_if:
            file_path = filedialog.askopenfilename(filetypes=[
                ("All files", "*.*"),
                ("Excel files", "*.xlsx;*.xls"),
                ("JSON files", "*.json"),
                ("Text files", "*.txt")])
            if file_path:
                self.load_if = file_path
                self.ui.tk_label_load_label.config(text=file_path)

                # 根据文件扩展名决定使用哪个库
                if file_path.endswith('.xlsx'):
                    wb = load_workbook(file_path)
                    sheet = wb.active  # 获取活动工作表
                    column_count = sheet.max_column
                elif file_path.endswith('.xls'):
                    wb = xlrd.open_workbook(file_path)
                    sheet = wb.sheet_by_index(0)  # 获取第一个工作表
                    column_count = sheet.ncols  # 获取列数
                elif file_path.endswith('.json'):
                    with open(file_path, 'r', encoding='utf-8') as file:
                        data = json.load(file)
                        if isinstance(data, dict) and data:
                            keys = data.keys()
                            column_options = ["数字", "字母", "密码", "自定义"] + list(keys)
                        else:
                            self.ui.tk_label_load_label.config(text="JSON文件不包含可解析的数据")
                            return
                elif file_path.endswith('.txt'):
                    with open(file_path, 'r', encoding='utf-8') as file:
                        lines = file.readlines()
                        if lines:
                            # 处理所有行，将每行的数据作为一个列表
                            column_options = ["数字", "字母", "密码", "自定义"] + [f"第{i+1}行" for i in range(len(lines))]
                        else:
                            self.ui.tk_label_load_label.config(text="TXT文件内容无效")
                            return
                else:
                    self.ui.tk_label_load_label.config(text="不支持的文件格式")
                    return

            cb = self.ui.tk_select_box_select_method  # 调用创建Combobox的方法
            cb['values'] = column_options  # 设置Combobox的选项
            cb.set(column_options[4] if len(column_options) > 4 else column_options[0])
        else:
            self.load_if = None
            self.ui.tk_label_load_label.config(text="目前读取内容：空")
            cb = self.ui.tk_select_box_select_method
            cb['values'] = ["数字","字母","密码", "自定义"]
            cb.set("数字")

    def handle_file_data(self, method, end_num, load_if):
        if load_if.endswith('.xlsx'):
            # 处理 .xlsx 文件
            wb = load_workbook(load_if, read_only=True)
            sheet = wb.active
            columns = int(method[1:-1])
            names_from_file = [
                sheet.cell(row=i, column=columns).value
                for i in range(1, end_num + 1)
                if sheet.cell(row=i, column=columns).value is not None
            ]
            wb.close()
        elif load_if.endswith('.xls'):
            # 处理 .xls 文件
            wb = xlrd.open_workbook(load_if)
            sheet = wb.sheet_by_index(0)  # 获取第一个工作表
            columns = int(method[1:-1]) - 1  # xlrd 的列索引从 0 开始
            names_from_file = [
                sheet.cell_value(rowx=i, colx=columns)
                for i in range(end_num)
                if sheet.cell_value(rowx=i, colx=columns) is not None
            ]
        elif load_if.endswith('.json'):
            with open(load_if, 'r', encoding='utf-8') as file:
                data = json.load(file)
                if method in data:
                    names_from_file = data[method][:end_num]
                else:
                    messagebox.showerror("错误", "指定的字段不存在于JSON文件中")
                    return
        elif load_if.endswith('.txt'):
            # 处理 .txt 文件
            with open(load_if, 'r', encoding='utf-8') as file:
                lines = file.readlines()

                # 处理行索引
                row_index = int(method[1:-1]) - 1  # 假设 method 是类似 "第2行" 的字符串
                if row_index < 0 or row_index >= len(lines):
                    messagebox.showerror("错误", "行索引超出范围")
                    return

                # 读取指定行的内容，并以空格分隔
                line_data = lines[row_index].strip().split()

                # 仅提取前 end_num 个字段，防止超出范围
                names_from_file = line_data[:end_num]

                if not names_from_file:
                    messagebox.showerror("错误", "指定的行为空或没有足够的数据")
                    return
        else:
            messagebox.showerror("错误", "不支持的文件格式")

        return names_from_file
